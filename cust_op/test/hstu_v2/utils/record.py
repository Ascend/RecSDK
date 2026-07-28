#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Copyright (c) Huawei Technologies Co., Ltd. 2026. All rights reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#    http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# ==============================================================================
import os
import csv
from typing import Dict, Any, Optional

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class Record:
    """测试结果记录类，用于将测试用例结果写入 Excel"""

    def __init__(self, output_path: str = "test_results.xlsx"):
        """初始化记录器

        Args:
            output_path: Excel 文件输出路径
        """
        self.output_path = output_path
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Test Results"

        # 表头样式
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")

        # 通过/失败样式
        self.pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # 边框样式
        self.border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )

        # 定义列头
        self.base_headers = [
            "Test Case Name",
            "Batch Size",
            "Head Num",
            "Head Dim QK",
            "Head Dim V",
            "Max SeqLen Q",
            "Max SeqLen K",
            "Has RAB",
            "Data Type",
            "Seed",
            "Is Metadata",
            "Overall Pass",
        ]

        self.grad_headers = ["DQ", "DK", "DV", "DRAB"]

        # 详细精度数据的列头模板
        self.detail_headers = ["Actual-FP32 Ref", "FP32 Ref", "Actual-Out Ref", "Passed"]

        # 序列长度统计的列头
        self.seq_stat_headers = [
            "SeqLen Q Mean",
            "SeqLen Q Max",
            "SeqLen Q Min",
            "SeqLen K Mean",
            "SeqLen K Max",
            "SeqLen K Min",
        ]

        # 初始化表头
        self._init_headers()

        self.current_row = 2  # 数据从第2行开始

    def record(
        self,
        case_name: str,
        params: Dict[str, Any],
        detail: Optional[Dict[str, Any]] = None,
        seq_stats: Optional[Dict[str, Any]] = None,
    ):
        """记录单个测试用例结果

        Args:
            case_name: 用例名称
            params: 用例参数字典，包含:
                ▪ batch_size

                ▪ head_num

                ▪ head_dim_qk

                ▪ head_dim_v

                ▪ max_seqlen_q

                ▪ max_seqlen_k

                ▪ has_rab

                ▪ data_type

                ▪ seed

            detail: 详细精度数据，由 Validator.backward_verify 返回的 detail 字典
            seq_stats: 序列长度统计数据，包含:
                ▪ seq_lens_q_mean, seq_lens_q_max, seq_lens_q_min

                ▪ seq_lens_k_mean, seq_lens_k_max, seq_lens_k_min

        """
        col = 1

        # 基础信息
        self.worksheet.cell(row=self.current_row, column=col, value=case_name)
        col += 1

        self.worksheet.cell(row=self.current_row, column=col, value=params.get("batch_size"))
        col += 1

        self.worksheet.cell(row=self.current_row, column=col, value=params.get("head_num"))
        col += 1

        self.worksheet.cell(row=self.current_row, column=col, value=params.get("head_dim_qk"))
        col += 1

        self.worksheet.cell(row=self.current_row, column=col, value=params.get("head_dim_v"))
        col += 1

        self.worksheet.cell(row=self.current_row, column=col, value=params.get("max_seqlen_q"))
        col += 1

        self.worksheet.cell(row=self.current_row, column=col, value=params.get("max_seqlen_k"))
        col += 1

        self.worksheet.cell(row=self.current_row, column=col, value=params.get("has_rab"))
        col += 1

        self.worksheet.cell(row=self.current_row, column=col, value=str(params.get("data_type")))
        col += 1

        self.worksheet.cell(row=self.current_row, column=col, value=params.get("seed"))
        col += 1

        self.worksheet.cell(row=self.current_row, column=col, value=params.get("is_metadata", False))
        col += 1

        # 判断是否通过
        passed = True
        if detail:
            for grad in self.grad_headers:
                grad_detail = detail.get(grad)
                if grad_detail is None:
                    continue
                if not self._check_pass(grad_detail):
                    passed = False
                    break

        # 写入 Overall Pass
        pass_cell = self.worksheet.cell(row=self.current_row, column=col, value=passed)
        pass_cell.fill = self.pass_fill if passed else self.fail_fill
        pass_cell.alignment = Alignment(horizontal="center", vertical="center")
        col += 1

        # 写入详细精度数据
        if detail:
            for grad in self.grad_headers:
                grad_detail = detail.get(grad)
                if grad_detail is None:
                    # RAB 可能为 None
                    for _ in self.detail_headers:
                        self.worksheet.cell(row=self.current_row, column=col, value="N/A")
                        col += 1
                    continue

                # 判断是否有详细精度数据
                has_detail_data = "actual-fp32_out_ref" in grad_detail

                if has_detail_data:
                    # 有详细精度数据 (pytorch_native_backend)
                    self.worksheet.cell(row=self.current_row, column=col, value=grad_detail.get("actual-fp32_out_ref"))
                    col += 1

                    self.worksheet.cell(row=self.current_row, column=col, value=grad_detail.get("fp32_out_ref"))
                    col += 1

                    self.worksheet.cell(row=self.current_row, column=col, value=grad_detail.get("actual-out_ref"))
                    col += 1
                else:
                    # 没有详细精度数据，只有 passed 状态 (ascend_native_backend)
                    self.worksheet.cell(row=self.current_row, column=col, value="N/A")
                    col += 1
                    self.worksheet.cell(row=self.current_row, column=col, value="N/A")
                    col += 1
                    self.worksheet.cell(row=self.current_row, column=col, value="N/A")
                    col += 1

                # 写入 passed 状态（✔ 或 ✗）
                grad_passed = grad_detail.get("passed", False)
                pass_symbol = "✔" if grad_passed else "✗"
                pass_cell = self.worksheet.cell(row=self.current_row, column=col, value=pass_symbol)
                pass_cell.fill = self.pass_fill if grad_passed else self.fail_fill
                pass_cell.alignment = Alignment(horizontal="center", vertical="center")
                col += 1
        else:
            # 没有详细数据，跳过
            for _ in self.grad_headers:
                for _ in self.detail_headers:
                    self.worksheet.cell(row=self.current_row, column=col, value="N/A")
                    col += 1

        # 写入序列长度统计数据
        if seq_stats:
            self.worksheet.cell(row=self.current_row, column=col, value=seq_stats.get("seq_lens_q_mean"))
            col += 1
            self.worksheet.cell(row=self.current_row, column=col, value=seq_stats.get("seq_lens_q_max"))
            col += 1
            self.worksheet.cell(row=self.current_row, column=col, value=seq_stats.get("seq_lens_q_min"))
            col += 1
            self.worksheet.cell(row=self.current_row, column=col, value=seq_stats.get("seq_lens_k_mean"))
            col += 1
            self.worksheet.cell(row=self.current_row, column=col, value=seq_stats.get("seq_lens_k_max"))
            col += 1
            self.worksheet.cell(row=self.current_row, column=col, value=seq_stats.get("seq_lens_k_min"))
            col += 1
        else:
            # 没有序列统计数据，跳过
            for _ in self.seq_stat_headers:
                self.worksheet.cell(row=self.current_row, column=col, value="N/A")
                col += 1

        # 添加边框
        for c in range(1, col):
            cell = self.worksheet.cell(row=self.current_row, column=c)
            cell.border = self.border

        self.current_row += 1

    def _check_pass(self, grad_detail: Dict[str, Any]) -> bool:
        """检查单个梯度是否通过验证"""
        # 优先使用 passed 字段
        if "passed" in grad_detail:
            return grad_detail["passed"]

        # 兼容旧逻辑
        actual_fp32_out_ref = grad_detail.get("actual-fp32_out_ref", 0)
        fp32_out_ref = grad_detail.get("fp32_out_ref", 0)
        multiplier = 5  # 与 Validator 中的 multiplier 一致

        # 使用与 Validator 相同的判断逻辑
        try_allclose = grad_detail.get("try_allclose", False)
        return (actual_fp32_out_ref <= multiplier * fp32_out_ref) or try_allclose

    def save(self):
        """保存 Excel 文件"""
        # 确保输出目录存在
        output_dir = os.path.dirname(self.output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)

        self.workbook.save(self.output_path)
        print(f"Test results saved to: {self.output_path}")

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.save()

    def _init_headers(self):
        """初始化表头"""
        headers = self.base_headers.copy()

        # 为每个梯度添加详细精度列
        for grad in self.grad_headers:
            for detail in self.detail_headers:
                headers.append(f"{grad}_{detail}")

        # 添加序列长度统计列头
        headers.extend(self.seq_stat_headers)

        # 写入表头
        for col, header in enumerate(headers, start=1):
            cell = self.worksheet.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

        # 设置列宽
        self.worksheet.column_dimensions['A'].width = 30  # Test Case Name
        for col in range(2, 12):
            self.worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
        # 序列统计列宽
        for col in range(12, 12 + len(self.seq_stat_headers)):
            self.worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14


class BenchmarkRecord:
    """Benchmark测试结果记录类，用于将测试用例输入写入CSV"""

    def __init__(self, output_path: str = "tmp_benchmark.csv"):
        """初始化记录器

        Args:
            output_path: CSV 文件输出路径
        """
        self.output_path = output_path
        self.headers = [
            'batch_size',
            'head_num',
            'max_seqlen_q',
            'max_seqlen_k',
            'head_dim_qk',
            'head_dim_v',
            'has_rab',
            'data_type',
            'seed',
            'is_metadata',
            'seq_lens_q_mean',
            'seq_lens_q_max',
            'seq_lens_q_min',
            'seq_lens_k_mean',
            'seq_lens_k_max',
            'seq_lens_k_min',
        ]
        self.current_row = 0
        # 初始化文件，先写入表头
        self._init_file()

    def _init_file(self):
        """初始化CSV文件"""
        if os.path.exists(self.output_path):
            os.remove(self.output_path)
        with open(self.output_path, 'w', newline='', encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(self.headers)

    def record(self, params: Dict[str, Any], seq_stats: Optional[Dict[str, Any]] = None):
        """记录单个测试用例输入参数

        Args:
            params: 用例参数字典，包含:
                ▪ batch_size

                ▪ head_num

                ▪ head_dim_qk

                ▪ head_dim_v

                ▪ max_seqlen_q

                ▪ max_seqlen_k

                ▪ has_rab

                ▪ data_type

                ▪ seed

            seq_stats: 序列长度统计数据，包含:
                ▪ seq_lens_q_mean, seq_lens_q_max, seq_lens_q_min

                ▪ seq_lens_k_mean, seq_lens_k_max, seq_lens_k_min

        """
        row = [
            params.get("batch_size"),
            params.get("head_num"),
            params.get("max_seqlen_q"),
            params.get("max_seqlen_k"),
            params.get("head_dim_qk"),
            params.get("head_dim_v"),
            params.get("has_rab"),
            str(params.get("data_type")).replace('torch.', ''),
            params.get("seed"),
            params.get("is_metadata", False),
        ]
        # 添加序列长度统计信息
        if seq_stats:
            row.extend(
                [
                    seq_stats.get("seq_lens_q_mean"),
                    seq_stats.get("seq_lens_q_max"),
                    seq_stats.get("seq_lens_q_min"),
                    seq_stats.get("seq_lens_k_mean"),
                    seq_stats.get("seq_lens_k_max"),
                    seq_stats.get("seq_lens_k_min"),
                ]
            )
        else:
            row.extend([None] * 6)

        with open(self.output_path, 'a', newline='', encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(row)
        self.current_row += 1

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass
